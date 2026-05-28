"""
Excel generator: structured event data → formatted Excel file.
- Sheet 1: Keywords (from default data)
- Sheet 2: Events List (extracted event data)
"""
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict


# ---- Keywords Sheet style constants ----
KW_HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="5B7FA0")
KW_HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
KW_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

KW_DATA_FONT = Font(name="微软雅黑", size=9, color="5B7FA0")
KW_DATA_FONT_BOLD = Font(name="微软雅黑", bold=True, size=9, color="5B7FA0")
KW_DATA_FONT_SMALL = Font(name="微软雅黑", size=8, color="5B7FA0")
KW_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
KW_DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

KW_SIDE_THIN = Side(style="thin", color="000000")
KW_SIDE_DOTTED = Side(style="dotted", color="000000")

KW_COL_WIDTHS = {
    "A": 12, "B": 24.7, "C": 13, "D": 18.7, "E": 36.9,
    "F": 18.7, "G": 13, "H": 12.8, "I": 36.9, "J": 3.2
}
KW_HEADER_ROW_HEIGHT = 15.6
KW_DATA_ROW_HEIGHT = 52.95
KW_SHORT_ROW_HEIGHT = 26.4


# ---- Events List Sheet style constants ----
EV_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
EV_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
EV_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

EV_DATA_FONT = Font(name="微软雅黑", size=9)
EV_LINK_FONT = Font(name="微软雅黑", size=9, color="0563C1", underline="single")

PRIORITY_HIGH_FONT = Font(
    name="微软雅黑", bold=True, italic=True, size=9,
    color="C00000", underline="double"
)

EV_SIDE_THIN = Side(style="thin", color="4472C4")
EV_SIDE_DOTTED = Side(style="dotted", color="4472C4")

EV_HEADER_BORDER_FIRST = Border(
    left=EV_SIDE_THIN, right=EV_SIDE_DOTTED,
    top=EV_SIDE_THIN, bottom=Side()
)
EV_HEADER_BORDER_MIDDLE = Border(
    left=EV_SIDE_DOTTED, right=EV_SIDE_DOTTED,
    top=EV_SIDE_THIN, bottom=Side()
)
EV_HEADER_BORDER_LAST = Border(
    left=EV_SIDE_DOTTED, right=EV_SIDE_THIN,
    top=EV_SIDE_THIN, bottom=Side()
)

EV_DATA_BORDER_FIRST = Border(
    left=EV_SIDE_THIN, right=EV_SIDE_DOTTED,
    top=EV_SIDE_THIN, bottom=Side()
)
EV_DATA_BORDER_MIDDLE = Border(
    left=EV_SIDE_DOTTED, right=EV_SIDE_DOTTED,
    top=EV_SIDE_THIN, bottom=Side()
)
EV_DATA_BORDER_LAST = Border(
    left=EV_SIDE_DOTTED, right=EV_SIDE_THIN,
    top=EV_SIDE_THIN, bottom=Side()
)

COLUMNS = [
    "No.", "事件类型", "Link", "Start Date", "End Date", "Priority",
    "Event Keywords", "Event English Keywords", "Event Description",
    "Headline", "备注/地点", "来源"
]

EV_COL_WIDTHS = {
    "A": 8.27, "B": 27.09, "C": 8.73, "D": 13, "E": 12.09,
    "F": 11, "G": 35, "H": 22.82, "I": 36.09,
    "J": 46, "K": 19.54, "L": 19.54
}


# ---- Keywords Sheet default data ----
DEFAULT_KEYWORDS_DATA = [
    ["Topic Group", "Topic", "Description", "Keyword 1", "Keyword 2",
     "Keyword 3", "Keyword 4", "Exclusionary", "发帖昵称包含", "备用关键词"],
    ["社会经济因素(Major Socioeconomic Events)",
     "大型会议和展览\nSignificant Conferences and Exhibitions",
     "在北京举办的大型展览会议",
     "北京", "会议,展览,博览会,峰会,交易会,服贸会,国际,论坛,年会,盛会,全球",
     "举办,举行", "", "地铁,公交,北京时间,动态,@北京",
     "中华人民共和国,新华网,新华社,央视新闻,人民日报,人民政府,北京日报,北京商报,BRTV,中国新闻,中国青年,中国日报", ""],
    ["",
     "体育赛事\nSports Events",
     "在北京举办的国际级别体育赛事及在北京举办的马拉松比赛",
     "北京", "世界田联,World Athletics,国际足联,国际篮联,国际泳联,国际体联,世界羽联,国际网联,国际自盟,国际排联",
     "", "", "地铁,公交,北京时间,动态,@北京",
     "中华人民共和国,新华网,新华社,央视新闻,人民日报,人民政府,北京日报,北京商报,BRTV,中国新闻,中国青年,中国日报", ",奥运会"],
    ["",
     "文娱活动\nCultural and Entertainment Activities",
     "在北京举办的电影节、音乐节、演唱会",
     "北京", "电影节,音乐节,艺术节,演唱会,",
     "", "", "地铁,公交,北京时间,动态,@北京",
     "中华人民共和国,新华网,新华社,央视新闻,人民日报,人民政府,北京日报,北京商报,BRTV,中国新闻,中国青年,中国日报", ""],
    ["",
     "节假日节庆\nHoliday Celebrations",
     "在北京举办的节假日的庆祝活动",
     "北京", "春节,清明,五一,中秋,国庆",
     '庙会,中秋灯会,"京彩"灯会,文化展演,演出季,非遗~10活动,文化~10活动,主题~10活动',
     "", "地铁,公交,北京时间,动态,@北京",
     "中华人民共和国,新华网,新华社,央视新闻,人民日报,人民政府,北京日报,北京商报,BRTV,中国新闻,中国青年,中国日报", ""],
    ["",
     "高级别政府会议\nHigh-profile Government Meetings",
     "两会、在北京举办的国际论坛、在京举办的其他政府会议",
     '全国人大~10会议,全国政协~10会议,中非合作论坛,全球共享发展行动论坛,"一带一路",北京,在京,习近平,李强,王毅',
     "全国人大~10会议,全国政协~10会议,高级别~10会议,中央~20会议,全国~20会议,亚太~20会议,中非合作论坛,",
     "", "", "地铁,公交,北京时间,动态,@北京,延安,上海,伦敦,巴黎,莫斯科,联合国",
     "中华人民共和国,新华网,新华社,央视新闻,人民日报,人民政府,北京日报,北京商报,BRTV,中国新闻,中国青年,中国日报", ""],
    ["自然因素 (Natural Factors)",
     "极端天气及自然灾害\nExtreme Weather & Natural Disasters",
     "全国范围内的极端天气及自然灾害导致的交通限制，如飞机停航列车停运\n极端天气 - 极端温度、极端降水、极端风速、极端湿度等",
     "高铁,铁路,民航,飞机,航班,机场,航站楼,火车,旅客列车",
     "热浪,寒潮,暴雨,干旱,大风,大雾,强风,低温冷害,霜冻,冻害,强降水,台风,龙卷风,热带气旋,雷暴,冰雹,强对流天气,",
     "列车停运,临时停运,停航,航班被取消,航班已取消,航班取消,航班将取消,取消部分航班",
     "国家防总,省防指,应急管理部,应急响应,中央气象台,蓝色预警,黄色预警,橙色预警,红色预警",
     "每日好物,好物推荐,新闻早报,每日新闻,早间新闻", "",
     "京哈铁路,京通铁路,京包铁路,京沪铁路,京九铁路,京广铁路,京张铁路,京雄城际铁路,京唐城际铁路,京滨城际铁路,"],
    ["政策因素 (Policy)",
     "中小学春秋假\nSpring and Autumn Breaks",
     "Short holidays scheduled in March–April ",
     "春假,春秋假,秋假安排,秋假期间,雪假,",
     "", "", "", "",
     "人民日报,央视新闻,新华社,中国日报,中国新闻网,北京发布,天津发布,上海发布,",
     ""],
    ["",
     "进出京政策\nEntry and Exit Beijing Policies",
     "政府部门发布的进出京政策",
     "进出京,出京政策,进京政策", "政策,管理",
     "", "", "", "", ""],
]


def _apply_keywords_styles(ws_kw):
    """Apply styles to Keywords Sheet."""
    for letter, width in KW_COL_WIDTHS.items():
        ws_kw.column_dimensions[letter].width = width

    ws_kw.row_dimensions[1].height = KW_HEADER_ROW_HEIGHT
    max_row = ws_kw.max_row
    for r in range(2, max_row + 1):
        ws_kw.row_dimensions[r].height = KW_DATA_ROW_HEIGHT
    ws_kw.row_dimensions[max_row].height = KW_SHORT_ROW_HEIGHT

    max_col = ws_kw.max_column
    for c in range(1, max_col + 1):
        cell = ws_kw.cell(row=1, column=c)
        cell.font = KW_HEADER_FONT
        cell.fill = KW_HEADER_FILL
        cell.alignment = KW_HEADER_ALIGNMENT
        left_side = KW_SIDE_THIN if c == 1 else Side()
        right_side = KW_SIDE_THIN if c == max_col else Side()
        cell.border = Border(
            left=left_side, right=right_side,
            top=KW_SIDE_THIN, bottom=Side()
        )

    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws_kw.cell(row=r, column=c)
            if c == 1:
                cell.font = KW_DATA_FONT_BOLD
                cell.alignment = KW_DATA_ALIGNMENT_CENTER
            elif c == 3:
                cell.font = KW_DATA_FONT_SMALL
                cell.alignment = KW_DATA_ALIGNMENT
            else:
                cell.font = KW_DATA_FONT
                cell.alignment = KW_DATA_ALIGNMENT

            left_side = KW_SIDE_THIN if c == 1 else Side()
            right_side = KW_SIDE_THIN if c == max_col else Side()
            cell.border = Border(
                left=left_side, right=right_side,
                top=Side(), bottom=KW_SIDE_DOTTED
            )

    if max_row >= 9:
        ws_kw.merge_cells("A2:A6")
        ws_kw.merge_cells("A8:A9")
    elif max_row >= 6:
        ws_kw.merge_cells("A2:A6")


def _apply_events_list_styles(ws, num_events: int):
    """Apply styles to Events List Sheet."""
    max_col = len(COLUMNS)

    for i, letter in enumerate([get_column_letter(c) for c in range(1, max_col + 1)]):
        width = EV_COL_WIDTHS.get(letter, 15)
        ws.column_dimensions[letter].width = width

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = EV_HEADER_FONT
        cell.fill = EV_HEADER_FILL
        cell.alignment = EV_HEADER_ALIGNMENT
        if c == 1:
            cell.border = EV_HEADER_BORDER_FIRST
        elif c == max_col:
            cell.border = EV_HEADER_BORDER_LAST
        else:
            cell.border = EV_HEADER_BORDER_MIDDLE

    for r in range(2, num_events + 2):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            col_name = COLUMNS[c - 1]

            if c == 1:
                cell.border = EV_DATA_BORDER_FIRST
            elif c == max_col:
                cell.border = EV_DATA_BORDER_LAST
            else:
                cell.border = EV_DATA_BORDER_MIDDLE

            cell.font = EV_DATA_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_name == "No.":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "事件类型":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_name == "Link":
                cell.font = EV_LINK_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_name in ("Start Date", "End Date"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Priority":
                if str(cell.value).lower() == "high":
                    cell.font = PRIORITY_HIGH_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Event Description":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_name in ("Headline", "备注/地点"):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_name == "来源":
                cell.alignment = Alignment(vertical="center")

        ws.row_dimensions[r].height = 40


def _parse_to_date(value):
    """Parse a string like '2026/5/15' or '2026-5-15' to a real date object.
    Returns the original value if parsing fails (so empty strings stay empty)."""
    if not value or value == "":
        return ""
    if isinstance(value, (date, datetime)):
        return value if isinstance(value, date) else value.date()
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%-m/%-d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try flexible parsing for "yyyy/m/d" without zero padding
    import re
    m = re.match(r"^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return s  # fallback: leave as text if unparseable


def build_excel(events: List[Dict], policy_count: int = 0) -> bytes:
    """
    Generate formatted Excel from event list, return bytes.

    Args:
        events: List of event dicts, keys match COLUMNS
        policy_count: Number of high-priority events (for compatibility)

    Returns:
        Excel file bytes
    """
    today = datetime.now()
    extraction_date = float(f"{today.month}.{today.day}")

    wb = Workbook()

    # ---- Sheet 1: Keywords ----
    ws_kw = wb.active
    ws_kw.title = "Keywords"

    for row_idx, row_data in enumerate(DEFAULT_KEYWORDS_DATA, 1):
        for col_idx, value in enumerate(row_data, 1):
            if isinstance(value, str) and "\\n" in value:
                value = value.replace("\\n", "\n")
            ws_kw.cell(row=row_idx, column=col_idx, value=value)

    _apply_keywords_styles(ws_kw)

    # ---- Sheet 2: Events List ----
    ws = wb.create_sheet(title="Events List")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for event_idx, event in enumerate(events):
        row = event_idx + 2

        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = event.get(col_name, "")

            if col_name == "No.":
                value = extraction_date
            elif col_name == "事件类型" and isinstance(value, str):
                value = value.replace("\\n", "\n")
            elif col_name in ("Start Date", "End Date"):
                value = _parse_to_date(value)

            cell = ws.cell(row=row, column=col_idx, value=value)

            if col_name == "Link" and value and str(value).startswith("http"):
                cell.hyperlink = str(value)

            if col_name in ("Start Date", "End Date") and isinstance(value, date):
                cell.number_format = "yyyy/m/d"

    _apply_events_list_styles(ws, len(events))

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(events) + 1}"

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
