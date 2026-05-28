"""
Beijing Major Events Information Extraction — CLI Tool

Extracts structured event data from files (Excel/CSV/TXT/JSON) using LLM,
generates a formatted Excel report.

Usage:
    python main.py file1.xlsx file2.xlsx --mode auto
    python main.py data.xlsx --mode spring_break
    python main.py report.xlsx --mode event_summary --category "大型会议和展览"
"""
import argparse
import json
import logging
import os
import sys
from datetime import datetime, date
from typing import List

from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:
    pass

from dotenv import load_dotenv

# Ensure src/ is discoverable when running from any working directory
_PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# Load .env before importing src modules (they read env vars at import time)
load_dotenv()

from src.extractor import (
    extract_events,
    extract_from_city_activity_list,
    guess_category_from_filename,
    count_policy_events,
    _deduplicate_events,
)
from src.preprocess import read_and_preprocess_file, pre_filter_raw_text
from src.excel_builder import build_excel
from src.file_utils import extract_filename_from_url
from src.concert_scraper import scrape_beijing_concerts

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# Max raw text chars before truncation
_MAX_RAW_TEXT_CHARS = 50_000
_SUPPORTED_EXT = ('.xlsx', '.xls', '.csv', '.txt', '.json', '.doc', '.docx', '.pdf')
_EVENT_COLUMNS = [
    "No.", "事件类型", "Link", "Start Date", "End Date", "Priority",
    "Event Keywords", "Event English Keywords", "Event Description",
    "Headline", "备注/地点", "来源"
]


def _format_excel_date(value) -> str:
    if isinstance(value, datetime):
        return f"{value.year}/{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.year}/{value.month}/{value.day}"
    return str(value).strip() if value is not None else ""


def _load_existing_events(output_path: str) -> List[dict]:
    if not os.path.exists(output_path):
        return []

    try:
        wb = load_workbook(output_path, data_only=True)
        if "Events List" not in wb.sheetnames:
            return []
        ws = wb["Events List"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        col_indexes = {col: headers.index(col) for col in _EVENT_COLUMNS if col in headers}
        events = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            event = {}
            has_content = False
            for col in _EVENT_COLUMNS:
                idx = col_indexes.get(col)
                value = row[idx] if idx is not None and idx < len(row) else ""
                if col in ("Start Date", "End Date"):
                    value = _format_excel_date(value)
                elif value is None:
                    value = ""
                else:
                    value = str(value).strip()
                if col != "No." and value:
                    has_content = True
                event[col] = value
            if has_content:
                events.append(event)

        logger.info(f"Loaded {len(events)} existing events from {output_path}")
        return events
    except Exception as e:
        logger.warning(f"Failed to load existing output for merge: {e}")
        return []


def _collect_input_files(input_dir: str) -> List[str]:
    if not os.path.isdir(input_dir):
        return []
    return [
        os.path.join(input_dir, fname)
        for fname in sorted(os.listdir(input_dir))
        if fname.lower().endswith(_SUPPORTED_EXT) and not fname.startswith('~')
    ]


def process_files(
    file_paths: List[str],
    mode: str = "auto",
    category: str = "",
    topic: str = "",
    extra_info: str = "",
    output_dir: str = "./output",
    include_concerts: bool = False,
    concert_force: bool = False,
    concert_start: str = "",
    concert_end: str = "",
    concert_venue_size: str = "medium",
) -> str:
    """
    Process one or more files, extract events, generate Excel.

    Returns:
        Path to generated Excel file.
    """
    all_events = []

    # Determine effective mode
    effective_mode = mode
    if effective_mode in ("auto", ""):
        if file_paths:
            detected_categories = set()
            for fp in file_paths:
                fname = os.path.basename(fp)
                cat = guess_category_from_filename(fname)
                if cat:
                    detected_categories.add(cat)

            if detected_categories == {"中小学春秋假"}:
                effective_mode = "spring_break"
                logger.info("Auto mode: all files are school break → spring_break mode")
            else:
                effective_mode = "event_summary"
                logger.info(
                    f"Auto mode: detected categories {detected_categories} "
                    f"→ event_summary mode (per-file processing)"
                )
        else:
            effective_mode = "concert_only" if include_concerts else "spring_break"

    logger.info(f"Effective mode: {effective_mode}")

    # Process each file
    for file_path in file_paths:
        fname = os.path.basename(file_path)

        if effective_mode == "spring_break":
            file_category = "中小学春秋假"
        else:
            file_category = guess_category_from_filename(fname) or category or "中小学春秋假"

        logger.info(f"Processing file: {fname}, category: {file_category}")

        # Structured document (city activity list) — direct read, no LLM extraction
        if file_category == "城市活动事件列表":
            try:
                events = extract_from_city_activity_list(
                    url_or_path=file_path,
                    category=file_category,
                )
                for e in events:
                    e["_source"] = "city_activity_list"
                all_events.extend(events)
                logger.info(f"File {fname}: extracted {len(events)} events from city activity list")
                continue
            except Exception as e:
                logger.warning(f"Failed to extract from city activity list {fname}: {e}")
                continue

        # Unstructured document: preprocess + LLM extraction
        text, _ = read_and_preprocess_file(file_path, effective_mode, file_category)

        if not text or text.startswith("["):
            logger.warning(f"Skipped file {fname}: {text[:100]}")
            continue

        # Truncate if too long
        if len(text) > _MAX_RAW_TEXT_CHARS:
            logger.warning(
                f"File {fname} text {len(text)} chars truncated to "
                f"{_MAX_RAW_TEXT_CHARS} chars"
            )
            text = text[:_MAX_RAW_TEXT_CHARS]

        if extra_info.strip():
            text += f"\n\n--- 用户补充说明 ---\n{extra_info.strip()}\n--- 说明结束 ---"

        try:
            events = extract_events(
                raw_text=text,
                mode=effective_mode,
                category=file_category,
                topic=topic,
                extra_info="",
            )
            for e in events:
                e["_source"] = "original_post"
            all_events.extend(events)
            logger.info(f"File {fname}: extracted {len(events)} events (category={file_category})")
        except Exception as e:
            logger.warning(f"Failed to extract from {fname}: {e}")
            continue

    # Scrape concerts if requested
    if include_concerts:
        try:
            print("\n🎤 正在采集演唱会信息...")
            concert_events = scrape_beijing_concerts(
                incremental=not concert_force,
                start_date=concert_start,
                end_date=concert_end,
                min_venue_size=concert_venue_size,
            )
            if concert_events:
                all_events.extend(concert_events)
                print(f"采集到 {len(concert_events)} 条演唱会信息")
            else:
                print("本次未采集到新的演唱会信息")
        except Exception as e:
            logger.warning(f"演唱会采集失败: {e}")
            print(f"⚠️ 演唱会采集失败: {e}")

    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y.%m.%d")
    display_filename = f"Travel_Facilitators_and_Hindrances_Events_{date_str}.xlsx"
    output_path = os.path.join(output_dir, display_filename)

    existing_events = _load_existing_events(output_path)
    if existing_events:
        all_events = existing_events + all_events
        logger.info(f"Merged with existing output: {len(existing_events)} existing events")

    if not all_events:
        print("未能从文档中提取到相关事件")
        return ""

    # Deduplicate + sort by date + renumber
    all_events = _deduplicate_events(all_events)

    def _parse_sort_date(event):
        d = str(event.get("Start Date", "")).strip()
        if not d:
            return datetime.max
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(d, fmt)
            except ValueError:
                continue
        return datetime.max

    all_events.sort(key=_parse_sort_date)

    for idx, event in enumerate(all_events, 1):
        event["No."] = idx

    logger.info(f"Total after dedup: {len(all_events)} events")

    # Generate Excel
    policy_count = count_policy_events(all_events)
    excel_bytes = build_excel(events=all_events, policy_count=policy_count)

    # Save to local file
    with open(output_path, "wb") as f:
        f.write(excel_bytes)

    logger.info(f"Excel saved to: {output_path}")

    # Print summary
    _print_summary(all_events, file_paths, effective_mode, policy_count, output_path)

    return output_path


def _print_summary(events, file_paths, effective_mode, policy_count, output_path):
    """Print markdown summary to console."""
    from src.extractor import get_category_type_name

    def _get_category_name(val):
        s = str(val)
        if "\n" in s:
            return s.split("\n", 1)[0].strip()
        if "\\n" in s:
            return s.split("\\n", 1)[0].strip()
        return s.strip()

    detected_categories = sorted(set(
        _get_category_name(e.get("事件类型", "")) for e in events
    ))
    detected_categories = [c for c in detected_categories if c]

    mode_name_map = {
        "spring_break": "春假提取模式 (spring_break)",
        "event_summary": "事件总结模式 (event_summary)",
        "auto": "自动分类模式 (auto)",
        "concert_only": "演唱会采集模式 (concert_only)",
    }
    mode_display = mode_name_map.get(effective_mode, f"事件总结模式 ({effective_mode})")

    total = len(events)

    # Category distribution
    dist_lines = []
    for cat in detected_categories:
        cnt = sum(1 for e in events if _get_category_name(e.get("事件类型", "")) == cat)
        pct = (cnt / total * 100) if total else 0
        dist_lines.append(f"{cat}：{cnt} 个（占 {pct:.1f}%）")
    category_dist = "\n".join(dist_lines)

    # File list
    file_names = "\n".join(os.path.basename(fp) for fp in file_paths[:50])

    # Important events (top 5 High Priority)
    high_events = [e for e in events if str(e.get("Priority", "")).strip().lower() == "high"]
    high_events_sorted = sorted(high_events, key=lambda x: str(x.get("Start Date", "")))
    important_lines = []
    for ev in high_events_sorted[:5]:
        headline = str(ev.get("Headline", "") or ev.get("Event Keywords", "")).strip()
        start = str(ev.get("Start Date", "")).strip()
        end = str(ev.get("End Date", "")).strip()
        location = str(ev.get("备注/地点", "") or ev.get("备注", "")).strip()

        date_str = f"{start} - {end}" if start and end and start != end else start
        location_short = location.split("，")[0].split(",")[0].split(" ")[0][:20] if location else ""
        parts = [p for p in [headline, date_str, location_short] if p]
        if parts:
            if location_short and len(parts) >= 2:
                event_line = f"{headline}（{date_str}，{location_short}）"
            else:
                event_line = f"{headline}（{date_str}）" if date_str else headline
            important_lines.append(event_line)

    summary = (
        f"📋 提取模式与检测到的事件类型\n"
        f"提取模式：{mode_display}\n"
        f"检测到的事件类型：{'、'.join(detected_categories)}\n\n"
        f"📊 事件统计\n"
        f"事件总数：{total} 个\n"
        f"高优先级事件 (High)：{policy_count} 个\n"
        f"普通事件：{total - policy_count} 个\n\n"
        f"📥 输出文件\n"
        f"{output_path}\n\n"
        f"📋 处理文件清单\n"
        f"本次处理了以下 {len(file_paths)} 个文件：\n"
        f"{file_names}\n\n"
        f"📊 事件类型分布\n"
        f"从 {total} 个提取的事件来看：\n"
        f"{category_dist}\n"
    )
    if important_lines:
        summary += f"\n重要事件包括：\n" + "\n".join(important_lines) + "\n"

    print(summary)


def main():
    parser = argparse.ArgumentParser(
        description="北京大事件信息提取 — 从文件中提取结构化事件数据并生成 Excel 报告"
    )
    parser.add_argument(
        "files",
        nargs="*",
        help="输入文件路径（支持 xlsx/xls/csv/txt/json，可同时传入多个文件）。留空则自动读取 --input-dir 下的所有文件",
    )
    parser.add_argument(
        "--input-dir",
        default="./input",
        help="输入目录（默认: ./input）。当未指定 files 参数时，自动读取此目录下所有支持的文件",
    )
    parser.add_argument(
        "--mode",
        choices=["auto", "spring_break", "event_summary"],
        default="auto",
        help="提取模式：auto（自动判断）、spring_break（中小学春秋假）、event_summary（事件总结）",
    )
    parser.add_argument(
        "--category",
        default="",
        help="手动指定事件分类（如『大型会议和展览』），留空则自动推断",
    )
    parser.add_argument(
        "--topic",
        default="",
        help="提取主题补充描述",
    )
    parser.add_argument(
        "--extra-info",
        default="",
        help="对提取要求的补充说明",
    )
    parser.add_argument(
        "--output-dir",
        default="./output",
        help="输出目录（默认: ./output）",
    )
    parser.add_argument(
        "--concert",
        action="store_true",
        help="同时从文化和旅游部网站采集北京地区演唱会信息（默认增量更新，跳过已采集的）",
    )
    parser.add_argument(
        "--concert-force",
        action="store_true",
        help="演唱会信息强制重新采集，不跳过已采集的 URL",
    )
    parser.add_argument(
        "--concert-start",
        default="",
        help="演唱会采集开始日期，格式 YYYY-MM-DD。默认过去2个月",
    )
    parser.add_argument(
        "--concert-end",
        default="",
        help="演唱会采集结束日期，格式 YYYY-MM-DD。默认未来2年",
    )
    parser.add_argument(
        "--concert-venue-size",
        choices=["small", "medium", "large"],
        default="medium",
        help="演唱会最小场馆规模：medium 默认排除小型 LiveHouse；large 只保留大型场馆；small 保留全部",
    )
    parser.add_argument(
        "--only-concert",
        action="store_true",
        help="只运行演唱会采集，跳过 input 文件处理。可与 --concert-force 搭配使用",
    )
    parser.add_argument(
        "--only-input",
        action="store_true",
        help="只处理 input 文件，不运行演唱会采集。若同时传入 --concert，本参数优先",
    )

    args = parser.parse_args()

    include_concerts = (args.concert or args.only_concert) and not args.only_input
    process_input = not args.only_concert

    # Collect file paths: from CLI args or from input directory
    file_paths = list(args.files) if process_input else []

    if process_input and not file_paths:
        input_dir = args.input_dir
        file_paths = _collect_input_files(input_dir)

        if file_paths:
            print(f"从 {input_dir}/ 读取到 {len(file_paths)} 个文件：")
            for fp in file_paths:
                print(f"  - {os.path.basename(fp)}")
        elif not include_concerts:
            if not os.path.isdir(input_dir):
                print(f"错误：未指定文件，且输入目录不存在: {input_dir}", file=sys.stderr)
                print(f"提示：请将文件放入 {input_dir}/ 目录，或在命令行直接指定文件路径", file=sys.stderr)
            else:
                print(f"错误：输入目录 {input_dir}/ 中没有支持的文件（xlsx/xls/csv/txt/json/doc/docx/pdf）", file=sys.stderr)
            sys.exit(1)

    # Validate files exist
    for fp in file_paths:
        if not fp.startswith(('http://', 'https://')) and not os.path.exists(fp):
            print(f"错误：文件不存在: {fp}", file=sys.stderr)
            sys.exit(1)

    # Validate API key (required if processing files, optional for concert-only mode)
    if file_paths and not os.getenv("OPENAI_API_KEY"):
        print("错误：未设置 OPENAI_API_KEY。请在 .env 文件或环境变量中配置。", file=sys.stderr)
        sys.exit(1)

    output_path = process_files(
        file_paths=file_paths,
        mode=args.mode,
        category=args.category,
        topic=args.topic,
        extra_info=args.extra_info,
        output_dir=args.output_dir,
        include_concerts=include_concerts,
        concert_force=args.concert_force,
        concert_start=args.concert_start,
        concert_end=args.concert_end,
        concert_venue_size=args.concert_venue_size,
    )

    if output_path:
        print(f"\n✅ 完成！Excel 文件已保存至: {output_path}")
    else:
        print("\n❌ 未能提取到任何事件")
        sys.exit(1)


if __name__ == "__main__":
    main()
